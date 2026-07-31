from pathlib import Path

from openpyxl import load_workbook

from doc_to_md.assets.naming import build_image_name
from doc_to_md.core.models import (
    ContentBlock,
    DocumentSection,
    ExtractedImage,
    ParsedDocument,
)
from doc_to_md.parsers.base_parser import BaseParser


class XlsxParser(BaseParser):
    def parse(self, file_path: Path, image_dir: Path) -> ParsedDocument:
        parsed = ParsedDocument(
            source_file=file_path.name,
            source_path=file_path,
            source_type="xlsx",
            title=file_path.stem,
            sections=[],
            images=[],
            warnings=[
                "Excel formulas may appear as formulas if cached calculated values are unavailable.",
                "Native Excel charts are not guaranteed to be extracted in this MVP.",
                "Embedded Excel images are extracted using openpyxl internal image access when available.",
            ],
            metadata={},
        )

        workbook = load_workbook(
            filename=str(file_path),
            data_only=False,
        )

        for sheet in workbook.worksheets:
            section = DocumentSection(
                heading=f"Sheet: {sheet.title}",
                level=1,
                blocks=[],
            )

            rows = self._sheet_to_rows(sheet)

            if rows:
                section.blocks.append(
                    ContentBlock(
                        type="table",
                        content=rows,
                    )
                )
            else:
                section.blocks.append(
                    ContentBlock(
                        type="warning",
                        content="This sheet is empty.",
                    )
                )

            sheet_images = self._extract_sheet_images(
                sheet=sheet,
                image_dir=image_dir,
            )

            parsed.images.extend(sheet_images)

            for image in sheet_images:
                section.blocks.append(
                    ContentBlock(
                        type="image",
                        content={
                            "alt": image.alt_text,
                            "path": image.relative_path,
                        },
                    )
                )

            parsed.sections.append(section)

        return parsed

    def _sheet_to_rows(self, sheet) -> list[list[str]]:
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0

        rows: list[list[str]] = []

        for row in sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ):
            values: list[str] = []

            for value in row:
                cell_value = "" if value is None else str(value).strip()
                values.append(cell_value)

            if any(values):
                rows.append(values)

        return self._trim_empty_columns(rows)

    def _trim_empty_columns(self, rows: list[list[str]]) -> list[list[str]]:
        if not rows:
            return []

        max_len = max(len(row) for row in rows)
        keep_indexes: list[int] = []

        for index in range(max_len):
            has_value = any(
                index < len(row) and bool(row[index])
                for row in rows
            )

            if has_value:
                keep_indexes.append(index)

        trimmed_rows: list[list[str]] = []

        for row in rows:
            trimmed_row = [
                row[index] if index < len(row) else ""
                for index in keep_indexes
            ]
            trimmed_rows.append(trimmed_row)

        return trimmed_rows

    def _extract_sheet_images(
        self,
        sheet,
        image_dir: Path,
    ) -> list[ExtractedImage]:
        extracted: list[ExtractedImage] = []
        images = getattr(sheet, "_images", [])

        for index, image in enumerate(images, start=1):
            try:
                image_bytes = image._data()
                extension = "png"

                if hasattr(image, "path") and "." in image.path:
                    extension = image.path.rsplit(".", 1)[-1].lower()

                filename = build_image_name(
                    prefix=f"sheet_{sheet.title}",
                    index=index,
                    extension=extension,
                )

                output_path = image_dir / filename
                output_path.write_bytes(image_bytes)

                extracted.append(
                    ExtractedImage(
                        filename=filename,
                        relative_path=f"./images/{filename}",
                        alt_text=f"Excel sheet {sheet.title} image {index}",
                    )
                )

            except Exception:
                continue

        return extracted
